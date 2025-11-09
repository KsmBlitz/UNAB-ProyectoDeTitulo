# app/routes/analytics.py
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional, Literal
from datetime import datetime, timedelta
from app.config import sensor_collection
from app.utils.dependencies import get_current_user
from app.services.cache import cache_service
from pydantic import BaseModel
import numpy as np
from scipy import stats
import io
import json
import logging
import traceback

logger = logging.getLogger(__name__)

# Librerías para exportación (instalar si no están)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

router = APIRouter()

# Modelos Pydantic
class CorrelationRequest(BaseModel):
    variable1: str
    variable2: str
    period: Literal["week", "month", "quarter"] = "month"

# Mapeo de períodos a días
PERIOD_DAYS = {
    "week": 7,
    "month": 30,
    "quarter": 90
}

@router.post("/correlation")
async def calculate_correlation(
    request: CorrelationRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Calcula la correlación de Pearson entre dos variables
    """
    # Intentar obtener desde caché
    cache_key = f"correlation_{request.variable1}_{request.variable2}_{request.period}"
    cached_data = await cache_service.get(cache_key)
    if cached_data:
        return cached_data
    
    try:
        days = PERIOD_DAYS.get(request.period, 30)
        start_date = datetime.utcnow() - timedelta(days=days)
        
        # Obtener datos desde MongoDB
        cursor = sensor_collection.find(
            {"ReadTime": {"$gte": start_date}},
            {"ReadTime": 1, request.variable1: 1, request.variable2: 1, "_id": 0}
        ).sort("ReadTime", 1)
        
        data = await cursor.to_list(length=10000)
        
        # Fallback: Si no hay suficientes datos en el período, usar todos los datos disponibles
        if len(data) < 2:
            logger.info(f"Insuficientes datos en período {request.period}, usando todos los datos disponibles")
            cursor = sensor_collection.find(
                {},
                {"ReadTime": 1, request.variable1: 1, request.variable2: 1, "_id": 0}
            ).sort("ReadTime", 1)
            data = await cursor.to_list(length=10000)
        
        if len(data) < 2:
            raise HTTPException(status_code=400, detail="No hay suficientes datos en la base de datos para calcular correlación")
        
        # Extraer valores
        var1_values = [d.get(request.variable1) for d in data if d.get(request.variable1) is not None]
        var2_values = [d.get(request.variable2) for d in data if d.get(request.variable2) is not None]
        
        # Asegurar que tengan la misma longitud
        min_length = min(len(var1_values), len(var2_values))
        var1_values = var1_values[:min_length]
        var2_values = var2_values[:min_length]
        
        if min_length < 2:
            raise HTTPException(status_code=400, detail="No hay suficientes datos válidos")
        
        # Verificar que haya variación en los datos
        var1_std = np.std(var1_values)
        var2_std = np.std(var2_values)
        
        if var1_std == 0 or var2_std == 0:
            # Si una variable es constante, la correlación no está definida
            result = {
                "coefficient": 0.0,
                "p_value": 1.0,
                "sample_size": min_length,
                "var1_mean": float(np.mean(var1_values)),
                "var2_mean": float(np.mean(var2_values)),
                "var1_std": float(var1_std),
                "var2_std": float(var2_std),
                "period": request.period,
                "note": "Una o ambas variables son constantes, correlación no definida"
            }
            await cache_service.set(cache_key, result, ttl=120)  # 2 minutos
            return result
        
        # Calcular correlación de Pearson
        correlation, p_value = stats.pearsonr(var1_values, var2_values)
        
        # Verificar si el resultado es válido (no NaN o Inf)
        if not np.isfinite(correlation):
            correlation = 0.0
        if not np.isfinite(p_value):
            p_value = 1.0
        
        result = {
            "coefficient": float(correlation),
            "p_value": float(p_value),
            "sample_size": min_length,
            "var1_mean": float(np.mean(var1_values)),
            "var2_mean": float(np.mean(var2_values)),
            "var1_std": float(var1_std),
            "var2_std": float(var2_std),
            "period": request.period
        }
        
        # Guardar en caché por 2 minutos
        await cache_service.set(cache_key, result, ttl=120)
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en correlación: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error al calcular correlación: {str(e)}")


@router.get("/anomalies")
async def detect_anomalies(
    period: Literal["week", "month", "quarter"] = "month",
    threshold: float = 2.5,  # Desviaciones estándar
    current_user: dict = Depends(get_current_user)
):
    """
    Detecta anomalías usando el método de desviaciones estándar (Z-score)
    """
    # Intentar obtener desde caché
    cache_key = f"anomalies_{period}_{threshold}"
    cached_data = await cache_service.get(cache_key)
    if cached_data:
        return cached_data
    
    try:
        days = PERIOD_DAYS.get(period, 30)
        start_date = datetime.utcnow() - timedelta(days=days)
        
        # Obtener datos
        cursor = sensor_collection.find(
            {"ReadTime": {"$gte": start_date}},
            {"ReadTime": 1, "pH_Value": 1, "Temperature": 1, "EC": 1, "reservoirId": 1, "_id": 0}
        ).sort("ReadTime", 1)
        
        data = await cursor.to_list(length=10000)
        
        # Fallback: Si no hay suficientes datos en el período, usar todos los datos disponibles
        if len(data) < 10:
            logger.info(f"Insuficientes datos en período {period} para anomalías, usando todos los datos disponibles")
            cursor = sensor_collection.find(
                {},
                {"ReadTime": 1, "pH_Value": 1, "Temperature": 1, "EC": 1, "reservoirId": 1, "_id": 0}
            ).sort("ReadTime", 1)
            data = await cursor.to_list(length=10000)
        
        if len(data) < 10:
            return {"anomalies": [], "message": "No hay suficientes datos en la base de datos"}
        
        anomalies = []
        
        # Analizar cada variable
        for variable in ["pH_Value", "Temperature", "EC"]:
            values = [d.get(variable) for d in data if d.get(variable) is not None]
            
            if len(values) < 10:
                continue
            
            mean = np.mean(values)
            std = np.std(values)
            
            # Detectar outliers usando Z-score
            for i, d in enumerate(data):
                value = d.get(variable)
                if value is None:
                    continue
                
                z_score = abs((value - mean) / std) if std > 0 else 0
                
                if z_score > threshold:
                    anomalies.append({
                        "timestamp": d.get("ReadTime").isoformat() if d.get("ReadTime") else None,
                        "sensor_type": variable,
                        "sensor_id": d.get("reservoirId", "unknown"),
                        "value": float(value),
                        "expected_range": f"{mean - 2*std:.2f} - {mean + 2*std:.2f}",
                        "z_score": float(z_score),
                        "severity": "Alta" if z_score > 3 else "Media",
                        "description": f"Valor fuera de rango normal (Z-score: {z_score:.2f})"
                    })
        
        # Ordenar por timestamp descendente
        anomalies.sort(key=lambda x: x["timestamp"], reverse=True)
        
        result = {
            "anomalies": anomalies[:50],  # Limitar a 50 anomalías más recientes
            "total": len(anomalies),
            "period": period
        }
        
        # Guardar en caché por 2 minutos
        await cache_service.set(cache_key, result, ttl=120)
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al detectar anomalías: {str(e)}")


@router.get("/predictions")
async def get_predictions(
    hours: int = Query(48, ge=1, le=168),  # Entre 1 y 168 horas (1 semana)
    current_user: dict = Depends(get_current_user)
):
    """
    Genera predicciones simples usando media móvil y tendencias lineales
    """
    # Intentar obtener desde caché
    cache_key = f"predictions_{hours}"
    cached_data = await cache_service.get(cache_key)
    if cached_data:
        return cached_data
    
    try:
        # Intentar obtener datos de las últimas 7 días
        start_date = datetime.utcnow() - timedelta(days=7)
        
        cursor = sensor_collection.find(
            {"ReadTime": {"$gte": start_date}},
            {"ReadTime": 1, "pH_Value": 1, "Temperature": 1, "EC": 1, "_id": 0}
        ).sort("ReadTime", 1)
        
        data = await cursor.to_list(length=10000)
        
        # Si no hay datos recientes, usar todos los datos disponibles
        if len(data) < 10:
            cursor = sensor_collection.find(
                {},
                {"ReadTime": 1, "pH_Value": 1, "Temperature": 1, "EC": 1, "_id": 0}
            ).sort("ReadTime", -1).limit(1000)
            
            data = await cursor.to_list(length=1000)
            
            if len(data) < 10:
                raise HTTPException(status_code=400, detail="No hay suficientes datos para predicción")
        
        predictions = {}
        
        for variable in ["pH_Value", "Temperature", "EC"]:
            values = [d.get(variable) for d in data if d.get(variable) is not None]
            
            if len(values) < 10:
                continue
            
            # Calcular tendencia usando regresión lineal simple
            x = np.arange(len(values))
            
            # Verificar si hay variación en los datos
            if np.std(values) == 0:
                # Valores constantes, no hay tendencia
                predictions[variable] = {
                    "predicted_value": float(values[-1]),
                    "current_value": float(values[-1]),
                    "trend": "Estable",
                    "slope": 0.0,
                    "confidence": "100%",
                    "hours_ahead": hours,
                    "note": "Valores constantes"
                }
                continue
            
            slope, intercept, r_value, p_value, std_err = stats.linregress(x, values)
            
            # Predecir valor futuro
            future_x = len(values) + (hours / 24) * (len(values) / 7)  # Proporción aproximada
            predicted_value = slope * future_x + intercept
            
            # Validar que los valores sean finitos
            if not np.isfinite(predicted_value):
                predicted_value = values[-1]
            if not np.isfinite(slope):
                slope = 0.0
            if not np.isfinite(r_value):
                r_value = 0.0
            
            # Determinar tendencia
            if abs(slope) < 0.01:
                trend = "Estable"
            elif slope > 0:
                trend = "Ascendente"
            else:
                trend = "Descendente"
            
            # Calcular confianza basada en R²
            confidence = min(100, max(0, r_value ** 2 * 100))
            
            predictions[variable] = {
                "predicted_value": float(predicted_value),
                "current_value": float(values[-1]),
                "trend": trend,
                "slope": float(slope),
                "confidence": f"{confidence:.0f}%",
                "hours_ahead": hours
            }
        
        # Guardar en caché por 5 minutos
        await cache_service.set(cache_key, predictions, ttl=300)
        return predictions
        
    except HTTPException:
        # Re-lanzar HTTPException sin modificar
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar predicciones: {str(e)}")


@router.get("/comparison")
async def get_historical_comparison(
    period: Literal["week", "month", "quarter"] = "month",
    current_user: dict = Depends(get_current_user)
):
    """
    Compara estadísticas del período actual con el período anterior
    """
    try:
        days = PERIOD_DAYS.get(period, 30)
        
        # Período actual
        current_start = datetime.utcnow() - timedelta(days=days)
        current_end = datetime.utcnow()
        
        # Período anterior
        previous_start = current_start - timedelta(days=days)
        previous_end = current_start
        
        # Función para obtener estadísticas
        async def get_stats(start, end):
            cursor = sensor_collection.find(
                {"ReadTime": {"$gte": start, "$lt": end}},
                {"pH_Value": 1, "Temperature": 1, "EC": 1, "_id": 0}
            )
            
            data = await cursor.to_list(length=10000)
            
            if not data:
                return None
            
            ph_values = [d.get("pH_Value") for d in data if d.get("pH_Value") is not None]
            temp_values = [d.get("Temperature") for d in data if d.get("Temperature") is not None]
            ec_values = [d.get("EC") for d in data if d.get("EC") is not None]
            
            return {
                "avgPH": float(np.mean(ph_values)) if ph_values else None,
                "avgTemp": float(np.mean(temp_values)) if temp_values else None,
                "avgEC": float(np.mean(ec_values)) if ec_values else None,
                "totalReadings": len(data),
                "start_date": start.isoformat(),
                "end_date": end.isoformat()
            }
        
        current_stats = await get_stats(current_start, current_end)
        previous_stats = await get_stats(previous_start, previous_end)
        
        # Fallback: Si no hay datos en los períodos especificados, usar todos los datos
        if current_stats is None and previous_stats is None:
            logger.info(f"No hay datos en período {period}, usando todos los datos disponibles")
            # Obtener todos los datos y dividirlos en dos mitades
            cursor = sensor_collection.find(
                {},
                {"ReadTime": 1, "pH_Value": 1, "Temperature": 1, "EC": 1, "_id": 0}
            ).sort("ReadTime", 1)
            
            all_data = await cursor.to_list(length=10000)
            
            if len(all_data) < 2:
                return {
                    "current": None,
                    "previous": None,
                    "period": period,
                    "message": "No hay suficientes datos en la base de datos"
                }
            
            # Dividir en dos mitades
            mid_point = len(all_data) // 2
            first_half = all_data[:mid_point]
            second_half = all_data[mid_point:]
            
            # Calcular estadísticas para cada mitad
            def calc_stats_from_list(data_list):
                if not data_list:
                    return None
                ph_values = [d.get("pH_Value") for d in data_list if d.get("pH_Value") is not None]
                temp_values = [d.get("Temperature") for d in data_list if d.get("Temperature") is not None]
                ec_values = [d.get("EC") for d in data_list if d.get("EC") is not None]
                
                return {
                    "avgPH": float(np.mean(ph_values)) if ph_values else None,
                    "avgTemp": float(np.mean(temp_values)) if temp_values else None,
                    "avgEC": float(np.mean(ec_values)) if ec_values else None,
                    "totalReadings": len(data_list),
                    "start_date": data_list[0].get("ReadTime").isoformat() if data_list else None,
                    "end_date": data_list[-1].get("ReadTime").isoformat() if data_list else None
                }
            
            previous_stats = calc_stats_from_list(first_half)
            current_stats = calc_stats_from_list(second_half)
        
        return {
            "current": current_stats,
            "previous": previous_stats,
            "period": period
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al comparar datos: {str(e)}")


@router.get("/export/excel")
async def export_excel(
    period: Literal["week", "month", "quarter"] = "month",
    current_user: dict = Depends(get_current_user)
):
    """
    Exporta datos analíticos a Excel
    """
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=501, detail="Librería openpyxl no instalada")
    
    try:
        days = PERIOD_DAYS.get(period, 30)
        start_date = datetime.utcnow() - timedelta(days=days)
        
        # Obtener datos
        cursor = sensor_collection.find(
            {"ReadTime": {"$gte": start_date}},
            {"ReadTime": 1, "pH_Value": 1, "Temperature": 1, "EC": 1, "reservoirId": 1, "_id": 0}
        ).sort("ReadTime", -1)
        
        data = await cursor.to_list(length=10000)
        
        # Fallback: Si no hay datos en el período, usar todos los datos disponibles
        if not data:
            logger.info(f"No hay datos en período {period} para Excel, usando todos los datos disponibles")
            cursor = sensor_collection.find(
                {},
                {"ReadTime": 1, "pH_Value": 1, "Temperature": 1, "EC": 1, "reservoirId": 1, "_id": 0}
            ).sort("ReadTime", -1)
            data = await cursor.to_list(length=10000)
        
        if not data:
            raise HTTPException(status_code=404, detail="No hay datos disponibles para exportar")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos de Sensores"
        
        # Headers
        headers = ["Fecha/Hora", "Sensor ID", "pH_Value", "Temperatura (°C)", "Electroconductividad"]
        ws.append(headers)
        
        # Estilo de headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Datos
        for item in data:
            ws.append([
                item.get("ReadTime").strftime("%Y-%m-%d %H:%M:%S") if item.get("ReadTime") else "",
                item.get("reservoirId", ""),
                item.get("pH_Value", ""),
                item.get("Temperature", ""),
                item.get("EC", "")
            ])
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        
        # Guardar en memoria
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=reporte-analitica-{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar Excel: {str(e)}")


@router.get("/export/pdf")
async def export_pdf(
    period: Literal["week", "month", "quarter"] = "month",
    current_user: dict = Depends(get_current_user)
):
    """
    Exporta datos analíticos a PDF
    """
    if not PDF_AVAILABLE:
        raise HTTPException(status_code=501, detail="Librería reportlab no instalada")
    
    try:
        days = PERIOD_DAYS.get(period, 30)
        start_date = datetime.utcnow() - timedelta(days=days)
        
        # Obtener estadísticas resumidas
        cursor = sensor_collection.find(
            {"ReadTime": {"$gte": start_date}},
            {"pH_Value": 1, "Temperature": 1, "EC": 1, "_id": 0}
        )
        
        data = await cursor.to_list(length=10000)
        
        # Fallback: Si no hay datos en el período, usar todos los datos disponibles
        if not data:
            logger.info(f"No hay datos en período {period} para PDF, usando todos los datos disponibles")
            cursor = sensor_collection.find(
                {},
                {"pH_Value": 1, "Temperature": 1, "EC": 1, "_id": 0}
            )
            data = await cursor.to_list(length=10000)
        
        if not data:
            raise HTTPException(status_code=404, detail="No hay datos disponibles para exportar")
        
        # Calcular estadísticas
        ph_values = [d.get("pH_Value") for d in data if d.get("pH_Value") is not None]
        temp_values = [d.get("Temperature") for d in data if d.get("Temperature") is not None]
        ec_values = [d.get("EC") for d in data if d.get("EC") is not None]
        
        # Crear PDF en memoria
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        
        # Contenido
        story = []
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph("<b>Reporte de Analítica Avanzada</b>", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 0.3*inch))
        
        # Información del reporte
        info_text = f"Período: {period.capitalize()}<br/>Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}<br/>Total de lecturas: {len(data)}"
        story.append(Paragraph(info_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Tabla de estadísticas
        table_data = [
            ['Variable', 'Promedio', 'Mínimo', 'Máximo', 'Desv. Est.'],
            [
                'pH',
                f"{np.mean(ph_values):.2f}" if ph_values else "N/A",
                f"{np.min(ph_values):.2f}" if ph_values else "N/A",
                f"{np.max(ph_values):.2f}" if ph_values else "N/A",
                f"{np.std(ph_values):.2f}" if ph_values else "N/A"
            ],
            [
                'Temperatura (°C)',
                f"{np.mean(temp_values):.2f}" if temp_values else "N/A",
                f"{np.min(temp_values):.2f}" if temp_values else "N/A",
                f"{np.max(temp_values):.2f}" if temp_values else "N/A",
                f"{np.std(temp_values):.2f}" if temp_values else "N/A"
            ],
            [
                'Electroconductividad',
                f"{np.mean(ec_values):.0f}" if ec_values else "N/A",
                f"{np.min(ec_values):.0f}" if ec_values else "N/A",
                f"{np.max(ec_values):.0f}" if ec_values else "N/A",
                f"{np.std(ec_values):.0f}" if ec_values else "N/A"
            ]
        ]
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Construir PDF
        doc.build(story)
        pdf_buffer.seek(0)
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=reporte-analitica-{datetime.now().strftime('%Y%m%d')}.pdf"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar PDF: {str(e)}")
